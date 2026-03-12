import logging
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.views.decorators.http import require_POST
from django.core.exceptions import PermissionDenied
from django.contrib.auth.views import redirect_to_login
from functools import wraps

logger = logging.getLogger(__name__)
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Count, Q, F, Value, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from datetime import datetime, timedelta
from accounts.models import CustomUser
from .models import Customer, InventoryItem, Expense, Payment, BillClaim, Sale, SaleItem, SalePayment, LedgerEntry, StockHistory
from django.core.paginator import Paginator
from .forms import CustomerForm, InventoryItemForm, ExpenseForm, PaymentForm, BillClaimForm, SaleForm, SaleItemForm, CombinedSaleItemForm, SalePaymentForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def _can_view_all_sales(user):
    return bool(user.is_superuser or getattr(user, 'is_manager', False))


def _visible_sales_queryset(user):
    qs = Sale.objects.select_related('customer').all()
    if _can_view_all_sales(user):
        return qs
    return qs.filter(created_by=user)


def _get_visible_sale_or_404(request, pk):
    return get_object_or_404(_visible_sales_queryset(request.user), pk=pk)

def is_manager(user):
    if not user.is_authenticated:
        return False
    return bool(
        user.is_superuser
        or getattr(user, 'is_manager', False)
        or user.groups.filter(name='Manager').exists()
    )

# Custom decorator: redirect unauthenticated to login, raise 403 for authenticated non-managers
def manager_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        user = request.user
        if not user.is_authenticated:
            return redirect_to_login(request.get_full_path(), login_url='/login/')
        if not is_manager(user):
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped


@login_required
def dashboard(request):
    """Dashboard with key metrics"""
    today = timezone.localdate()

    def _machine_label_from_description(text: str) -> str:
        # Common patterns in this project: free-text, "Machine: <name> - <details>", or multi-line.
        if not text:
            return ''
        label = str(text).strip().splitlines()[0].strip()
        # Strip common prefixes
        for prefix in ('machine:', 'Machine:', 'MACHINE:'):
            if label.startswith(prefix):
                label = label[len(prefix):].strip()
                break
        # If it looks like "Name - details" or "Name — details", keep only the name.
        for sep in (' - ', ' — ', ' – '):
            if sep in label:
                label = label.split(sep, 1)[0].strip()
                break
        return label

    # Top selling products (all time) by quantity, across inventory + machine items
    inventory_top = (
        SaleItem.objects.filter(sale__status='finalized', item_type='inventory', inventory_item__isnull=False)
        .values('inventory_item__part_name')
        .annotate(total_qty=Sum('quantity'))
    )

    machine_qty_by_label = {}
    for desc, qty in (
        SaleItem.objects.filter(sale__status='finalized', item_type='non_inventory')
        .exclude(description='')
        .values_list('description', 'quantity')
    ):
        label = _machine_label_from_description(desc)
        if not label:
            continue
        machine_qty_by_label[label] = (machine_qty_by_label.get(label) or 0) + (qty or 0)

    top_products = [
        {'label': row['inventory_item__part_name'], 'item_type': 'inventory', 'total_qty': row['total_qty'] or 0}
        for row in inventory_top
        if row.get('inventory_item__part_name')
    ] + [
        {'label': label, 'item_type': 'machine', 'total_qty': total_qty}
        for label, total_qty in machine_qty_by_label.items()
    ]
    top_products = sorted(top_products, key=lambda r: r['total_qty'], reverse=True)[:10]
    top_products_max_qty = max([p['total_qty'] for p in top_products], default=0)

    context = {
        'total_employees': CustomUser.objects.filter(status='active').count(),
        'total_customers': Customer.objects.filter(status='active').count(),
        'total_inventory_items': InventoryItem.objects.count(),
        'low_stock_items': InventoryItem.objects.filter(quantity__lte=F('minimum_stock')).count(),
        # Ensure mixed int/decimal multiplication resolves to decimal via ExpressionWrapper
        'total_inventory_value': InventoryItem.objects.aggregate(
            total=Sum(
                ExpressionWrapper(
                    F('quantity') * Coalesce(F('unit_price'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))),
                    output_field=DecimalField(max_digits=12, decimal_places=2)
                )
            )
        )['total'] or 0,
        # Sale metrics (replacing legacy Payment metrics)
        'pending_sales': Sale.objects.filter(status='draft').count(),
        'finalized_sales': Sale.objects.filter(status='finalized').count(),
        'today_sales_total': Sale.objects.filter(status='finalized', finalized_at__date=today).aggregate(
            total=Sum('total_amount')
        )['total'] or 0,
        'top_products': top_products,
        'top_products_max_qty': top_products_max_qty,
        'total_sales_due': Sale.objects.filter(status='finalized').annotate(
            balance=F('total_amount') - Coalesce(Sum('payments__amount'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=2)))
        ).aggregate(total=Sum('balance'))['total'] or 0,
        'monthly_expenses': Expense.objects.filter(
            date__month=datetime.now().month,
            date__year=datetime.now().year
        ).aggregate(total=Sum('amount'))['total'] or 0,
        'recent_expenses': Expense.objects.all()[:5],
        'recent_sales': Sale.objects.select_related('customer').all()[:5],
        'pending_bill_claims': BillClaim.objects.filter(status='pending').count(),
        'low_stock_alerts': InventoryItem.objects.filter(
            quantity__lte=F('minimum_stock')
        )[:5],
    }
    return render(request, 'core/dashboard.html', context)


# Customer Views
@login_required
@permission_required('core.view_customer', raise_exception=True)
def customer_list(request):
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    qs = Customer.objects.all()
    if query:
        qs = qs.filter(
            Q(name__icontains=query) |
            Q(customer_id__icontains=query) |
            Q(company__icontains=query) |
            Q(phone__icontains=query)
        )
    if status_filter:
        qs = qs.filter(status=status_filter)
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/customer_list.html', {
        'customers': page_obj.object_list,
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
    })


@login_required
@permission_required('core.add_customer', raise_exception=True)
def customer_add(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer added successfully!')
            return redirect('customer_list')
    else:
        form = CustomerForm()
    return render(request, 'core/customer_form.html', {'form': form, 'title': 'Add Customer'})


@login_required
@permission_required('core.change_customer', raise_exception=True)
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer updated successfully!')
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'core/customer_form.html', {'form': form, 'title': 'Edit Customer'})


@login_required
@permission_required('core.delete_customer', raise_exception=True)
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        messages.success(request, 'Customer deleted successfully!')
        return redirect('customer_list')
    return render(request, 'core/confirm_delete.html', {'object': customer, 'type': 'Customer'})


@login_required
@permission_required('core.view_customer', raise_exception=True)
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    sales_qs = customer.sales.filter(status='finalized').select_related().prefetch_related('items__inventory_item', 'payments').order_by('-created_at')
    
    # Calculate aggregate totals for all finalized orders
    total_amount = sales_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    total_paid = sum(sale.total_paid for sale in sales_qs)
    total_due = total_amount - total_paid
    
    paginator = Paginator(sales_qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    context = {
        'customer': customer,
        'sales': page_obj.object_list,
        'page_obj': page_obj,
        'title': 'Customer Details',
        'total_amount': total_amount,
        'total_paid': total_paid,
        'total_due': total_due,
        'add_payment_form': SalePaymentForm(initial={'payment_date': timezone.localdate()}),
        'customer_payment_error': request.session.pop('customer_payment_error', ''),
    }
    return render(request, 'core/customer_detail.html', context)


@login_required
@permission_required('core.add_salepayment', raise_exception=True)
@require_POST
def customer_add_payment(request, customer_id):
    form = SalePaymentForm(request.POST)

    if not form.is_valid():
        errors = []
        for field, errs in form.errors.items():
            label = form.fields.get(field).label if field in form.fields else field
            for err in errs:
                errors.append(f"{label}: {err}")
        customer = get_object_or_404(Customer, pk=customer_id)
        customer_url = reverse('customer_detail', kwargs={'pk': customer.pk})
        request.session['customer_payment_error'] = ' '.join(errors) if errors else 'Please correct the highlighted fields.'
        return redirect(f"{customer_url}?open_customer_payment=1#orders")

    cleaned = form.cleaned_data
    try:
        incoming_amount = Decimal(cleaned.get('amount') or 0)
    except (TypeError, InvalidOperation):
        incoming_amount = Decimal('0')

    customer = get_object_or_404(Customer, pk=customer_id)
    customer_url = reverse('customer_detail', kwargs={'pk': customer.pk})

    if incoming_amount <= 0:
        request.session['customer_payment_error'] = 'Amount must be greater than zero.'
        return redirect(f"{customer_url}?open_customer_payment=1#orders")

    base_notes = (cleaned.get('notes') or '').strip()
    note_prefix = 'Payment via customer account - FIFO allocation'
    composed_notes = note_prefix if not base_notes else f"{note_prefix} | {base_notes}"

    with transaction.atomic():
        # Serialize customer-level allocations per customer in a Postgres-safe way.
        customer = Customer.objects.select_for_update().get(pk=customer.pk)

        candidate_sales = list(
            customer.sales.filter(status='finalized')
            .order_by('finalized_at', 'id')
            .select_for_update()
            .prefetch_related('payments')
        )

        due_sales = []
        total_due = Decimal('0')
        for sale in candidate_sales:
            paid_amount = sum((payment.amount for payment in sale.payments.all()), Decimal('0'))
            due_amount = (sale.total_amount or Decimal('0')) - paid_amount
            if due_amount > 0:
                sale.due_amount = due_amount
                due_sales.append(sale)
                total_due += due_amount

        if total_due <= 0:
            request.session['customer_payment_error'] = 'No unpaid finalized sales found for this customer.'
            return redirect(f"{customer_url}#orders")

        if incoming_amount > total_due:
            request.session['customer_payment_error'] = f'Payment exceeds customer total due (Tk {total_due:.2f}).'
            return redirect(f"{customer_url}?open_customer_payment=1#orders")

        remaining = incoming_amount
        allocations = []
        payment_date = cleaned.get('payment_date')
        method = cleaned.get('method')

        for sale in due_sales:
            if remaining <= 0:
                break
            allocation = min(remaining, sale.due_amount)
            if allocation <= 0:
                continue

            payment = SalePayment.objects.create(
                sale=sale,
                amount=allocation,
                payment_date=payment_date,
                method=method,
                notes=composed_notes,
            )
            try:
                LedgerEntry.objects.update_or_create(
                    source='sale_payment',
                    reference=payment.receipt_number,
                    defaults={
                        'entry_type': 'credit',
                        'description': f"Payment for {sale.sale_number}",
                        'amount': payment.amount,
                    },
                )
            except Exception:
                logger.exception('Non-blocking ledger write failure for SalePayment receipt=%s', payment.receipt_number)

            allocations.append((sale.sale_number, allocation))
            remaining -= allocation

    allocation_text = ', '.join([f"{sale_no}: Tk {amount:.2f}" for sale_no, amount in allocations])
    messages.success(
        request,
        f"Payment recorded (Tk {incoming_amount:.2f}) across {len(allocations)} sale(s). Allocation: {allocation_text}",
    )
    return redirect(f"{customer_url}#orders")


# Inventory Views
@login_required
@permission_required('core.view_inventoryitem', raise_exception=True)
def inventory_list(request):
    query = request.GET.get('q', '')
    category_filter = request.GET.get('category', '')
    low_stock = request.GET.get('low_stock', '')
    qs = InventoryItem.objects.all()
    if query:
        qs = qs.filter(
            Q(part_name__icontains=query) |
            Q(part_code__icontains=query) |
            Q(category__icontains=query)
        )
    if category_filter:
        qs = qs.filter(category__icontains=category_filter)
    if low_stock:
        qs = qs.filter(quantity__lte=F('minimum_stock'))
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/inventory_list.html', {
        'items': page_obj.object_list,
        'page_obj': page_obj,
        'query': query,
        'category_filter': category_filter,
        'low_stock': low_stock,
    })


@login_required
@permission_required('core.add_inventoryitem', raise_exception=True)
def inventory_add(request):
    if request.method == 'POST':
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            item = form.save()
            if item.quantity > 0:
                StockHistory.objects.create(
                    item=item,
                    transaction_type='in',
                    quantity=item.quantity,
                    previous_quantity=0,
                    new_quantity=item.quantity,
                    box_quantity=0,
                    previous_box_quantity=0,
                    new_box_quantity=0,
                    reason='Initial stock',
                    created_by=request.user.username,
                )
            messages.success(request, 'Inventory item added successfully!')
            return redirect('inventory_list')
    else:
        form = InventoryItemForm()
    return render(request, 'core/inventory_form.html', {'form': form, 'title': 'Add Inventory Item'})


@login_required
@permission_required('core.change_inventoryitem', raise_exception=True)
def inventory_edit(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    previous_quantity = item.quantity
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item)
        if form.is_valid():
            item = form.save()
            new_quantity = item.quantity
            if new_quantity != previous_quantity:
                diff = new_quantity - previous_quantity
                if diff > 0:
                    tx_type = 'in'
                    reason = 'Stock added via edit'
                else:
                    tx_type = 'adjustment'
                    reason = 'Stock adjusted via edit'
                StockHistory.objects.create(
                    item=item,
                    transaction_type=tx_type,
                    quantity=abs(diff),
                    previous_quantity=previous_quantity,
                    new_quantity=new_quantity,
                    box_quantity=0,
                    previous_box_quantity=0,
                    new_box_quantity=0,
                    reason=reason,
                    created_by=request.user.username,
                )
            messages.success(request, 'Inventory item updated successfully!')
            return redirect('inventory_list')
    else:
        form = InventoryItemForm(instance=item)
    return render(request, 'core/inventory_form.html', {'form': form, 'title': 'Edit Inventory Item'})


@login_required
@permission_required('core.delete_inventoryitem', raise_exception=True)
def inventory_delete(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    if request.method == 'POST':
        item.delete()
        messages.success(request, 'Inventory item deleted successfully!')
        return redirect('inventory_list')
    return render(request, 'core/confirm_delete.html', {'object': item, 'type': 'Inventory Item'})


@login_required
@permission_required('core.view_inventoryitem', raise_exception=True)
def inventory_stock_history(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    history = item.stock_history.select_related().all()
    paginator = Paginator(history, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/inventory_stock_history.html', {
        'item': item,
        'history': page_obj.object_list,
        'page_obj': page_obj,
    })


# Expense Views
@login_required
@permission_required('core.view_expense', raise_exception=True)
def expense_list(request):
    query = request.GET.get('q', '')
    category_filter = request.GET.get('category', '')
    month_filter = request.GET.get('month', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    qs = Expense.objects.all()
    if query:
        qs = qs.filter(
            Q(description__icontains=query) |
            Q(paid_to__icontains=query) |
            Q(receipt_number__icontains=query)
        )
    if category_filter:
        qs = qs.filter(category=category_filter)
    # Month filter (accept both YYYY-MM from input type="month" and MM-YYYY)
    if month_filter and not (start_date or end_date):  # prioritize explicit date range
        try:
            parts = month_filter.split('-')
            if len(parts) == 2:
                # Detect ordering: if first part length == 4 treat as year-month
                if len(parts[0]) == 4:
                    year = int(parts[0])
                    month = int(parts[1])
                else:
                    month = int(parts[0])
                    year = int(parts[1])
                qs = qs.filter(date__year=year, date__month=month)
        except Exception:
            logger.exception('Failed to parse month_filter: %s', month_filter)

    # Specific date or date range filtering
    from django.utils.dateparse import parse_date
    sd = parse_date(start_date) if start_date else None
    ed = parse_date(end_date) if end_date else None
    if sd and ed:
        qs = qs.filter(date__range=(sd, ed))
    elif sd:
        qs = qs.filter(date=sd)
    elif ed:
        qs = qs.filter(date=ed)
    total_expenses = qs.aggregate(total=Sum('amount'))['total'] or 0
    credit_total = LedgerEntry.objects.filter(entry_type='credit').aggregate(total=Sum('amount'))['total'] or 0
    debit_total = LedgerEntry.objects.filter(entry_type='debit').aggregate(total=Sum('amount'))['total'] or 0
    current_balance = (credit_total or 0) - (debit_total or 0)
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/expense_list.html', {
        'expenses': page_obj.object_list,
        'page_obj': page_obj,
        'query': query,
        'category_filter': category_filter,
        'month_filter': month_filter,
        'start_date': start_date,
        'end_date': end_date,
        'total_expenses': total_expenses,
        'balance': current_balance,
    })


@login_required
@permission_required('core.add_expense', raise_exception=True)
def expense_add(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            exp = form.save()  # Ledger entry now created exclusively by post_save signal
            messages.success(request, 'Expense added successfully!')
            return redirect('expense_list')
    else:
        form = ExpenseForm()
    return render(request, 'core/expense_form.html', {'form': form, 'title': 'Add Expense'})


@login_required
@manager_required
def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expense updated successfully!')
            return redirect('expense_list')
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'core/expense_form.html', {'form': form, 'title': 'Edit Expense'})


@login_required
@manager_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Expense deleted successfully!')
        return redirect('expense_list')
    return render(request, 'core/confirm_delete.html', {'object': expense, 'type': 'Expense'})


@login_required
@permission_required('core.view_expense', raise_exception=True)
def expense_detail(request, pk):
    """View expense details"""
    expense = get_object_or_404(Expense, pk=pk)
    
    # Check if this expense was created from a bill claim
    related_claim = None
    related_claim_submitter = None
    related_claim_approver = None
    related_claim_approval_date = None
    try:
        related_claim = BillClaim.objects.get(expense=expense)
    except BillClaim.DoesNotExist:
        pass
    else:
        submitter_name = related_claim.submitter.get_full_name().strip()
        if not submitter_name:
            submitter_name = related_claim.submitter.username
        related_claim_submitter = submitter_name

        if related_claim.approved_by:
            approver_name = related_claim.approved_by.get_full_name().strip()
            if not approver_name:
                approver_name = related_claim.approved_by.username
            related_claim_approver = approver_name
        related_claim_approval_date = related_claim.approval_date
    
    context = {
        'expense': expense,
        'related_claim': related_claim,
        'related_claim_submitter': related_claim_submitter,
        'related_claim_approver': related_claim_approver,
        'related_claim_approval_date': related_claim_approval_date,
        'title': 'Expense Details'
    }
    return render(request, 'core/expense_detail.html', context)


# Payment Views - DEPRECATED: Legacy payment tracking system, replaced by SalePayment
# All payment processing now goes through SalePayment which is directly linked to Sales
# Keeping code commented for reference and potential data migration needs

# @login_required
# @user_passes_test(lambda u: u.is_superuser)
# def payment_list(request):
#     query = request.GET.get('q', '')
#     status_filter = request.GET.get('status', '')
#     payment_type = request.GET.get('payment_type', '')
#     qs = Payment.objects.all()
#     if query:
#         qs = qs.filter(
#             Q(invoice_number__icontains=query) |
#             Q(customer__name__icontains=query) |
#             Q(customer__customer_id__icontains=query)
#         )
#     if status_filter:
#         qs = qs.filter(status=status_filter)
#     if payment_type:
#         qs = qs.filter(payment_type=payment_type)
#     total_pending = qs.filter(status__in=['pending', 'overdue']).aggregate(
#         total=Sum(F('total_amount') - F('paid_amount'))
#     )['total'] or 0
#     paginator = Paginator(qs, 10)
#     page_obj = paginator.get_page(request.GET.get('page'))
#     return render(request, 'core/payment_list.html', {
#         'payments': page_obj.object_list,
#         'page_obj': page_obj,
#         'query': query,
#         'status_filter': status_filter,
#         'payment_type': payment_type,
#         'total_pending': total_pending
#     })


# @login_required
# @user_passes_test(lambda u: u.is_superuser)
# def payment_add(request):
#     if request.method == 'POST':
#         form = PaymentForm(request.POST)
#         if form.is_valid():
#             form.save()
#             messages.success(request, 'Payment added successfully!')
#             return redirect('payment_list')
#     else:
#         form = PaymentForm()
#     return render(request, 'core/payment_form.html', {'form': form, 'title': 'Add Payment'})


# @login_required
# @user_passes_test(lambda u: u.is_superuser)
# def payment_edit(request, pk):
#     payment = get_object_or_404(Payment, pk=pk)
#     if request.method == 'POST':
#         form = PaymentForm(request.POST, instance=payment)
#         if form.is_valid():
#             form.save()
#             messages.success(request, 'Payment updated successfully!')
#             return redirect('payment_list')
#     else:
#         form = PaymentForm(instance=payment)
#     return render(request, 'core/payment_form.html', {'form': form, 'title': 'Edit Payment'})


# @login_required
# @user_passes_test(lambda u: u.is_superuser)
# def payment_delete(request, pk):
#     payment = get_object_or_404(Payment, pk=pk)
#     if request.method == 'POST':
#         payment.delete()
#         messages.success(request, 'Payment deleted successfully!')
#         return redirect('payment_list')
#     return render(request, 'core/confirm_delete.html', {'object': payment, 'type': 'Payment'})


# Bill Claim Views
@login_required
@permission_required('core.submit_bill', raise_exception=True)
def submit_bill_claim(request):
    """Employee submits a new bill claim"""
    if request.method == 'POST':
        form = BillClaimForm(request.POST, request.FILES)
        if form.is_valid():
            bill_claim = form.save(commit=False)
            bill_claim.submitter = request.user
            bill_claim.status = 'pending'
            bill_claim.save()
            messages.success(request, 'Bill claim submitted successfully!')
            return redirect('my_bill_claims')
    else:
        form = BillClaimForm()

    return render(request, 'core/bill_claim_form.html', {'form': form, 'title': 'Submit Bill Claim'})


@login_required
def my_bill_claims(request):
    """Employee views their own submitted claims"""
    status_filter = request.GET.get('status', '')
    bill_claims = BillClaim.objects.filter(submitter=request.user).order_by('-created_at')

    if status_filter:
        bill_claims = bill_claims.filter(status=status_filter)

    context = {
        'bill_claims': bill_claims,
        'status_filter': status_filter,
        'title': 'My Bill Claims',
        'is_manager_view': False,
    }
    return render(request, 'core/bill_claim_list.html', context)


@login_required
@manager_required
def list_bill_claims(request):
    """Manager views all claims from all employees"""
    status_filter = request.GET.get('status', '')
    query = request.GET.get('q', '')
    
    bill_claims = BillClaim.objects.all().order_by('-created_at')

    if status_filter:
        bill_claims = bill_claims.filter(status=status_filter)
    
    if query:
        bill_claims = bill_claims.filter(
            Q(submitter__username__icontains=query) |
            Q(submitter__first_name__icontains=query) |
            Q(submitter__last_name__icontains=query) |
            Q(description__icontains=query)
        )

    # Calculate totals
    total_pending = bill_claims.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
    total_approved = bill_claims.filter(status='approved').aggregate(total=Sum('amount'))['total'] or 0
    total_rejected = bill_claims.filter(status='rejected').aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'bill_claims': bill_claims,
        'status_filter': status_filter,
        'query': query,
        'title': 'All Bill Claims',
        'is_manager_view': True,
        'total_pending': total_pending,
        'total_approved': total_approved,
        'total_rejected': total_rejected,
    }
    return render(request, 'core/bill_claim_list.html', context)


@login_required
@manager_required
def approve_bill_claim(request, pk):
    """Manager approves a bill claim and creates expense"""
    bill_claim = get_object_or_404(BillClaim, pk=pk)
    
    if bill_claim.status != 'pending':
        messages.warning(request, 'This claim has already been processed.')
        return redirect('list_bill_claims')

    if request.method == 'POST':
        from django.utils import timezone
        
        bill_claim.status = 'approved'
        bill_claim.approved_by = request.user
        bill_claim.approval_date = timezone.now().date()
        
        # Get employee name (fallback to username if no full name)
        employee_name = bill_claim.submitter.get_full_name().strip()
        if not employee_name:
            employee_name = bill_claim.submitter.username
        
        # Create a new Expense entry
        expense = Expense.objects.create(
            date=bill_claim.bill_date,
            category='other',
            description=f"Bill Claim by {employee_name}: {bill_claim.description}",
            amount=bill_claim.amount,
            paid_to=employee_name,
            payment_method='bank_transfer',
            notes=f"Approved bill claim (ID: {bill_claim.pk})",
        )
        bill_claim.expense = expense
        bill_claim.save()
        
        messages.success(request, f'Bill claim approved and expense created! Amount: ৳ {bill_claim.amount}')
        return redirect('list_bill_claims')

    context = {
        'bill_claim': bill_claim,
        'title': 'Approve Bill Claim',
        'action': 'approve',
    }
    return render(request, 'core/bill_claim_review.html', context)


@login_required
@manager_required
def reject_bill_claim(request, pk):
    """Manager rejects a bill claim"""
    bill_claim = get_object_or_404(BillClaim, pk=pk)
    
    if bill_claim.status != 'pending':
        messages.warning(request, 'This claim has already been processed.')
        return redirect('list_bill_claims')

    if request.method == 'POST':
        from django.utils import timezone
        
        bill_claim.status = 'rejected'
        bill_claim.approved_by = request.user
        bill_claim.approval_date = timezone.now().date()
        bill_claim.save()
        
        messages.warning(request, f'Bill claim rejected for {bill_claim.submitter.get_full_name()}.')
        return redirect('list_bill_claims')

    context = {
        'bill_claim': bill_claim,
        'title': 'Reject Bill Claim',
        'action': 'reject',
    }
    return render(request, 'core/bill_claim_review.html', context)


# Reports
@login_required
@manager_required
def reports(request):
    """Reports page with various statistics"""
    # Monthly expense breakdown
    monthly_expenses = Expense.objects.filter(
        date__year=datetime.now().year
    ).values('category').annotate(total=Sum('amount')).order_by('-total')
    
    # Payment statistics
    payment_stats = {
        'pending': Payment.objects.filter(status='pending').count(),
        'completed': Payment.objects.filter(status='completed').count(),
        'overdue': Payment.objects.filter(status='overdue').count(),
    }
    
    # Current balance from ledger
    credit_total = LedgerEntry.objects.filter(entry_type='credit').aggregate(total=Sum('amount'))['total'] or 0
    debit_total = LedgerEntry.objects.filter(entry_type='debit').aggregate(total=Sum('amount'))['total'] or 0
    current_balance = (credit_total or 0) - (debit_total or 0)

    context = {
        'monthly_expenses': monthly_expenses,
        'payment_stats': payment_stats,
        'balance': current_balance,
    }
    return render(request, 'core/reports.html', context)


@login_required
@manager_required
def ledger(request):
    """Simple ledger listing showing credits and debits with current balance."""
    qs = LedgerEntry.objects.all().order_by('-timestamp')
    credit_total = qs.filter(entry_type='credit').aggregate(total=Sum('amount'))['total'] or 0
    debit_total = qs.filter(entry_type='debit').aggregate(total=Sum('amount'))['total'] or 0
    current_balance = (credit_total or 0) - (debit_total or 0)
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    context = {
        'entries': page_obj.object_list,
        'page_obj': page_obj,
        'balance': current_balance,
        'credit_total': credit_total,
        'debit_total': debit_total,
    }
    return render(request, 'core/ledger.html', context)


@login_required
@manager_required
def export_excel(request):
    """Export all data to Excel"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Export Employees
    ws_employees = wb.create_sheet("Employees")
    ws_employees.append(['ID', 'Name', 'Position', 'Department', 'Salary', 'Status', 'Join Date'])
    for emp in CustomUser.objects.all():
        ws_employees.append([
            emp.employee_id, emp.get_full_name(), emp.position, emp.department, 
            float(emp.salary) if emp.salary else 0, emp.status, emp.join_date
        ])
    
    # Export Customers
    ws_customers = wb.create_sheet("Customers")
    ws_customers.append(['ID', 'Name', 'Company', 'Phone', 'City', 'Status'])
    for cust in Customer.objects.all():
        ws_customers.append([
            cust.customer_id, cust.name, cust.company, cust.phone, cust.city, cust.status
        ])
    
    # Export Inventory
    ws_inventory = wb.create_sheet("Inventory")
    ws_inventory.append(['Product Code', 'Product Name', 'Category', 'Quantity', 'Unit Price', 'Total Value'])
    for item in InventoryItem.objects.all():
        ws_inventory.append([
            item.part_code, item.part_name, item.category, item.quantity,
            float(item.unit_price or 0), float(item.total_value)
        ])
    
    # Export Expenses
    ws_expenses = wb.create_sheet("Expenses")
    ws_expenses.append(['Date', 'Category', 'Description', 'Amount', 'Paid To'])
    for exp in Expense.objects.all():
        ws_expenses.append([
            exp.date, exp.category, exp.description, float(exp.amount), exp.paid_to
        ])
    
    # Export Payments
    ws_payments = wb.create_sheet("Payments")
    ws_payments.append(['Invoice', 'Customer', 'Type', 'Total Amount', 'Paid Amount', 'Remaining', 'Status'])
    for pay in Payment.objects.all():
        ws_payments.append([
            pay.invoice_number, pay.customer.name, pay.payment_type, 
            float(pay.total_amount), float(pay.paid_amount), 
            float(pay.remaining_amount), pay.status
        ])
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=org_management_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


@login_required
@manager_required
def customer_report_excel(request):
    """Export customer financial summary to Excel with Total, Paid, and Due amounts"""
    from django.db.models import Sum, F, DecimalField
    from django.db.models.functions import Coalesce
    
    # Get all customers with their sales aggregates
    customers = Customer.objects.all().order_by('customer_id')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Report"
    
    # Header styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    headers = ['Name', 'Company', 'Phone', 'Total', 'Paid', 'Due']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data rows
    total_amount_sum = 0
    total_paid_sum = 0
    total_due_sum = 0
    
    for customer in customers:
        # Get all finalized sales for this customer
        sales = customer.sales.filter(status='finalized')
        
        # Calculate totals
        total_amount = sales.aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Calculate total paid from all sale payments
        total_paid = 0
        for sale in sales:
            total_paid += sale.total_paid
        
        balance_due = total_amount - total_paid
        
        # Add to grand totals
        total_amount_sum += total_amount
        total_paid_sum += total_paid
        total_due_sum += balance_due
        
        # Add row
        ws.append([
            customer.name,
            customer.company or '-',
            customer.phone,
            float(total_amount),
            float(total_paid),
            float(balance_due)
        ])
    
    # Add totals row
    ws.append([])
    total_row = ws.max_row + 1
    ws.append(['TOTAL', '', '', float(total_amount_sum), float(total_paid_sum), float(total_due_sum)])
    
    # Style totals row
    total_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    total_font = Font(bold=True, size=11)
    for cell in ws[total_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center' if cell.column <= 3 else 'right')
    
    # Format currency columns (D, E, F) with alignment and number format
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=customer_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


# =========================
# Sales Views (minimal UI)
# =========================

from django.forms import formset_factory


def _collect_form_errors(*, sale_form=None, item_formset=None, payment_form=None, limit=4):
    parts = []

    def _flatten_errors(errors_dict):
        msgs = []
        for field, errs in errors_dict.items():
            if field == '__all__':
                label = 'General'
            else:
                label = str(field).replace('_', ' ').title()
            for e in errs:
                msgs.append(f"{label}: {e}")
        return msgs

    if sale_form is not None and getattr(sale_form, 'errors', None):
        msgs = _flatten_errors(sale_form.errors)
        if msgs:
            parts.append("Customer - " + "; ".join(msgs[:limit]))

    if item_formset is not None:
        msgs = []
        try:
            non_form = list(item_formset.non_form_errors())
            msgs.extend(non_form)
        except Exception:
            logger.exception('Error reading non_form_errors from item_formset')
        for i, f in enumerate(getattr(item_formset, 'forms', []) or []):
            if getattr(f, 'errors', None):
                row_msgs = _flatten_errors(f.errors)
                for m in row_msgs:
                    msgs.append(f"Item {i + 1} - {m}")
            if len(msgs) >= limit:
                break
        if msgs:
            parts.append("Items - " + "; ".join(msgs[:limit]))

    if payment_form is not None and getattr(payment_form, 'errors', None):
        msgs = _flatten_errors(payment_form.errors)
        if msgs:
            parts.append("Payment - " + "; ".join(msgs[:limit]))

    return parts


@login_required
@permission_required('core.view_sale', raise_exception=True)
def sale_list(request):
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    item_type = request.GET.get('item_type', '')  # 'inventory' or 'machine'
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    selected_user_id = request.GET.get('user_id', '').strip()
    can_filter_by_user = _can_view_all_sales(request.user)
    qs = _visible_sales_queryset(request.user).order_by('-created_at')
    if query:
        qs = qs.filter(
            Q(sale_number__icontains=query) |
            Q(customer__name__icontains=query) |
            Q(customer__customer_id__icontains=query)
        )
    if status:
        qs = qs.filter(status=status)
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            qs = qs.filter(created_at__date__gte=start_date_obj)
        except ValueError:
            start_date = ''
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            qs = qs.filter(created_at__date__lte=end_date_obj)
        except ValueError:
            end_date = ''
    if item_type in ['inventory', 'machine']:
        # Map 'machine' to non_inventory sale items
        mapped = 'non_inventory' if item_type == 'machine' else 'inventory'
        qs = qs.filter(items__item_type=mapped).distinct()
    if can_filter_by_user and selected_user_id:
        try:
            qs = qs.filter(created_by_id=int(selected_user_id))
        except (TypeError, ValueError):
            selected_user_id = ''
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    # Sales overview metrics
    # Business rule: exclude Draft & Quotation from aggregate totals and dues
    totals_qs = qs.exclude(status__in=['draft', 'quote'])  # after filters
    total_sales_amount = totals_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    total_paid_amount = (
        totals_qs.annotate(paid=Sum('payments__amount')).aggregate(total=Sum('paid'))['total'] or 0
    )
    total_due_amount = (total_sales_amount or 0) - (total_paid_amount or 0)
    sales_users = []
    if can_filter_by_user:
        sales_users = CustomUser.objects.filter(status='active').order_by('first_name', 'username')
    return render(request, 'core/sale_list.html', {
        'sales': page_obj.object_list,
        'page_obj': page_obj,
        'query': query,
        'status': status,
        'item_type': item_type,
        'start_date': start_date,
        'end_date': end_date,
        'sales_users': sales_users,
        'selected_user_id': selected_user_id,
        'can_filter_by_user': can_filter_by_user,
        'total_sales_amount': total_sales_amount,
        'total_paid_amount': total_paid_amount,
        'total_due_amount': total_due_amount,
    })


@login_required
@permission_required('core.add_sale', raise_exception=True)
def sale_create_unified(request):
    """Unified sale creation: customer + multiple items + optional initial payment.
    Steps 1-6 implementation (no finalize toggle yet).
    """
    SaleItemFormSet = formset_factory(SaleItemForm, extra=1, can_delete=True)
    if request.method == 'POST':
        # Preprocess: copy machine quantity/unit price into hidden fields; description already bound in template
        post_data = request.POST.copy()
        try:
            total_forms = int(post_data.get('items-TOTAL_FORMS', '0'))
        except ValueError:
            total_forms = 0
        
        for i in range(total_forms):
            t = post_data.get(f'items-{i}-item_type')
            
            if t == 'non_inventory':
                desc_key = f'items-{i}-description'
                qty_key = f'items-{i}-quantity'
                qty_machine_key = f'items-{i}-quantity_machine'
                price_key = f'items-{i}-unit_price'
                price_machine_key = f'items-{i}-unit_price_machine'
                
                # Description already posted via textarea bound to formset
                
                # Copy machine quantity
                qty_val = post_data.get(qty_key, '').strip()
                qty_machine_val = post_data.get(qty_machine_key, '').strip()
                if not qty_val and qty_machine_val:
                    post_data[qty_key] = qty_machine_val
                
                # Copy machine price
                price_val = post_data.get(price_key, '').strip()
                price_machine_val = post_data.get(price_machine_key, '').strip()
                if not price_val and price_machine_val:
                    post_data[price_key] = price_machine_val

        sale_form = SaleForm(post_data)
        item_formset = SaleItemFormSet(post_data, prefix='items')
        payment_form = SalePaymentForm(request.POST, prefix='pay')
        
        valid = True
        if not sale_form.is_valid():
            valid = False
        # Note: We build items directly from POST, not from formset
        if payment_form.is_valid():
            pass
        else:
            # Allow empty payment section (all fields blank) -> treat as no payment
            # If user entered amount but form invalid, block with a specific message
            amt_raw = request.POST.get('pay-amount')
            if amt_raw:
                payment_form.add_error('amount', 'Enter a valid payment amount.')
                valid = False
        # Basic item presence validation - build items directly from POST data
        cleaned_items = []
        
        for idx in range(total_forms):
            item_type = post_data.get(f'items-{idx}-item_type', '').strip()
            
            # Skip completely blank rows
            if not item_type:
                continue
                
            if item_type == 'inventory':
                inv_id = post_data.get(f'items-{idx}-inventory_item')
                qty = post_data.get(f'items-{idx}-quantity', post_data.get(f'items-{idx}-quantity_inv', '1'))
                price = post_data.get(f'items-{idx}-unit_price', post_data.get(f'items-{idx}-unit_price_inv', '0'))
                
                if not inv_id:
                    messages.error(request, 'Inventory item must be selected')
                    valid = False
                    continue
                    
                try:
                    inv_item = InventoryItem.objects.get(pk=inv_id)
                    from decimal import Decimal as _D, InvalidOperation
                    try:
                        qty_dec = _D(str(qty)) if qty else _D('1')
                    except InvalidOperation:
                        qty_dec = _D('1')
                    try:
                        price_dec = _D(str(price)) if price else None
                    except InvalidOperation:
                        price_dec = None
                    cleaned_items.append({
                        'item_type': 'inventory',
                        'inventory_item': inv_item,
                        'description': inv_item.name,
                        'quantity': qty_dec,
                        'unit_price': price_dec if price_dec is not None else inv_item.unit_price
                    })
                except InventoryItem.DoesNotExist:
                    messages.error(request, 'Invalid inventory item')
                    valid = False
                    
            elif item_type == 'non_inventory':
                machine_desc = post_data.get(f'items-{idx}-description', '').strip()
                qty = post_data.get(f'items-{idx}-quantity', post_data.get(f'items-{idx}-quantity_machine', '1'))
                price = post_data.get(f'items-{idx}-unit_price', post_data.get(f'items-{idx}-unit_price_machine', '0'))
                
                if not machine_desc:
                    messages.error(request, 'Machine details are required')
                    valid = False
                    continue
                
                from decimal import Decimal as _D, InvalidOperation
                try:
                    qty_dec = _D(str(qty)) if qty else _D('1')
                except InvalidOperation:
                    qty_dec = _D('1')
                try:
                    price_dec = _D(str(price)) if price else _D('0')
                except InvalidOperation:
                    price_dec = _D('0')
                item_dict = {
                    'item_type': 'non_inventory',
                    'inventory_item': None,
                    'description': machine_desc,
                    'quantity': qty_dec,
                    'unit_price': price_dec
                }
                cleaned_items.append(item_dict)
        if not cleaned_items:
            valid = False
            messages.error(request, 'Add at least one valid item.')
        # Compute total from items (unit_price fallback to inventory price if missing)
        from decimal import Decimal as _D
        total_amount = _D('0')
        if cleaned_items:
            for cd in cleaned_items:
                unit_price = cd.get('unit_price') or _D('0')
                # If inventory and unit_price zero, fallback
                if cd.get('item_type') == 'inventory' and cd.get('inventory_item') and (unit_price is None or unit_price <= 0):
                    unit_price = _D(str(cd['inventory_item'].unit_price or 0))
                qty = cd.get('quantity') or _D('0')
                total_amount += _D(str(unit_price)) * _D(str(qty))
        # Payment validation (if provided)
        payment_amount = None
        if payment_form.is_valid():
            payment_amount = payment_form.cleaned_data.get('amount')
            if payment_amount is not None and payment_amount < 0:
                payment_form.add_error('amount', 'Payment cannot be negative.')
                valid = False
            elif cleaned_items and payment_amount and payment_amount > 0:
                if payment_amount > total_amount:
                    payment_form.add_error('amount', 'Payment exceeds total.')
                    valid = False
        if valid:
            from django.db import transaction
            with transaction.atomic():
                sale = sale_form.save(commit=False)
                sale.created_by = request.user
                # status stays draft for now
                sale.save()
                for cd in cleaned_items:
                    unit_price = cd.get('unit_price') or 0
                    if cd.get('item_type') == 'inventory' and cd.get('inventory_item') and (unit_price is None or unit_price <= 0):
                        unit_price = cd['inventory_item'].unit_price or 0
                    SaleItem.objects.create(
                        sale=sale,
                        item_type=cd['item_type'],
                        inventory_item=cd.get('inventory_item') if cd['item_type'] == 'inventory' else None,
                        description=(cd.get('description') or '') if cd['item_type'] != 'inventory' else (
                            f"{cd['inventory_item'].part_name} ({cd['inventory_item'].part_code})" if cd.get('inventory_item') else ''
                        ),
                        quantity=cd.get('quantity') or 0,
                        boxes=cd.get('boxes') or 0,
                        unit_price=unit_price,
                    )
                sale.recalc_total(save=True)
                payment = None
                if payment_amount and payment_amount > 0:
                    payment = payment_form.save(commit=False)
                    payment.sale = sale
                    payment.save()
                messages.success(request, 'Sale created successfully.')
                if payment:
                    messages.success(request, f'Payment recorded (Receipt {payment.receipt_number}).')
                return redirect('sale_detail', pk=sale.pk)
        else:
            parts = _collect_form_errors(sale_form=sale_form, payment_form=payment_form)
            messages.error(request, " | ".join(parts) if parts else 'Please correct the highlighted fields.')
    else:
        sale_form = SaleForm()
        item_formset = formset_factory(SaleItemForm, extra=1, can_delete=True)(prefix='items')
        payment_form = SalePaymentForm(prefix='pay')
    # Inventory price data for JS
    inv_prices = {str(i.id): float(i.unit_price or 0) for i in InventoryItem.objects.all()}
    return render(request, 'core/sale_create_unified.html', {
        'sale_form': sale_form,
        'item_formset': item_formset,
        'payment_form': payment_form,
        'inventory_prices': inv_prices,
        'title': 'Create Sale'
    })


@login_required
@permission_required('core.add_sale', raise_exception=True)
def sale_create(request):
    """Unified multi-item + optional payment create (invoice mode)."""
    SaleItemFormSet = formset_factory(SaleItemForm, extra=1, can_delete=True)
    if request.method == 'POST':
        sale_form = SaleForm(request.POST)
        item_formset = SaleItemFormSet(request.POST, prefix='items')
        payment_form = SalePaymentForm(request.POST, prefix='pay')
        valid = True
        if not sale_form.is_valid():
            valid = False
        if not item_formset.is_valid():
            valid = False
        if not payment_form.is_valid():
            amt_raw = request.POST.get('pay-amount')
            if amt_raw:
                payment_form.add_error('amount', 'Enter a valid payment amount.')
                valid = False
        cleaned_items = []
        if item_formset.is_valid():
            for f in item_formset.forms:
                if f.cleaned_data.get('DELETE'): continue
                cd = f.cleaned_data
                if not cd.get('item_type') and not cd.get('inventory_item') and not cd.get('description'):
                    continue
                if not cd.get('item_type'):
                    f.add_error('item_type', 'Select type'); valid = False; continue
                if cd.get('item_type') == 'non_inventory' and not cd.get('description'):
                    f.add_error('description', 'Description required'); valid = False; continue
                if cd.get('item_type') == 'inventory' and not cd.get('inventory_item'):
                    f.add_error('inventory_item', 'Choose inventory item'); valid = False; continue
                cleaned_items.append(cd)
        if not cleaned_items:
            messages.error(request, 'Add at least one valid item.'); valid = False
        from decimal import Decimal as _D
        total_amount = _D('0')
        for cd in cleaned_items:
            unit_price = cd.get('unit_price') or _D('0')
            if cd.get('item_type') == 'inventory' and cd.get('inventory_item') and (unit_price is None or unit_price <= 0):
                unit_price = _D(str(cd['inventory_item'].unit_price or 0))
            qty = cd.get('quantity') or _D('0')
            total_amount += _D(str(unit_price)) * _D(str(qty))
        payment_amount = None
        if payment_form.is_valid():
            payment_amount = payment_form.cleaned_data.get('amount')
            if payment_amount is not None and payment_amount < 0:
                payment_form.add_error('amount', 'Payment cannot be negative.'); valid = False
            elif cleaned_items and payment_amount and payment_amount > total_amount:
                payment_form.add_error('amount', 'Payment exceeds total.'); valid = False
        if valid:
            from django.db import transaction
            with transaction.atomic():
                sale = sale_form.save(commit=False)
                sale.created_by = request.user
                sale.save()
                for cd in cleaned_items:
                    unit_price = cd.get('unit_price') or 0
                    if cd.get('item_type') == 'inventory' and cd.get('inventory_item') and (unit_price is None or unit_price <= 0):
                        unit_price = cd['inventory_item'].unit_price or 0
                    SaleItem.objects.create(
                        sale=sale,
                        item_type=cd['item_type'],
                        inventory_item=cd.get('inventory_item') if cd['item_type'] == 'inventory' else None,
                        description=(cd.get('description') or '') if cd['item_type'] != 'inventory' else (
                            f"{cd['inventory_item'].part_name} ({cd['inventory_item'].part_code})" if cd.get('inventory_item') else ''
                        ),
                        quantity=cd.get('quantity') or 0,
                        boxes=cd.get('boxes') or 0,
                        unit_price=unit_price,
                    )
                sale.recalc_total(save=True)
                payment = None
                if payment_amount and payment_amount > 0:
                    payment = payment_form.save(commit=False)
                    payment.sale = sale
                    payment.save()
                messages.success(request, 'Sale created successfully.')
                if payment:
                    messages.success(request, f'Payment recorded (Receipt {payment.receipt_number}).')
                return redirect('sale_detail', pk=sale.pk)
        else:
            parts = _collect_form_errors(sale_form=sale_form, item_formset=item_formset, payment_form=payment_form)
            messages.error(request, " | ".join(parts) if parts else 'Please correct the highlighted fields.')
    else:
        sale_form = SaleForm()
        item_formset = formset_factory(SaleItemForm, extra=1, can_delete=True)(prefix='items')
        payment_form = SalePaymentForm(prefix='pay')
    inventory_items = InventoryItem.objects.all()
    inv_prices = {str(i.id): float(i.unit_price or 0) for i in inventory_items}
    return render(request, 'core/sale_create_unified.html', {
        'sale_form': sale_form,
        'item_formset': item_formset,
        'payment_form': payment_form,
        'inventory_items': inventory_items,
        'inventory_prices': inv_prices,
        'title': 'Create Sale'
    })


@login_required
@permission_required('core.add_sale', raise_exception=True)
def sale_quote_create(request):
    """Unified multi-item + optional payment create (quotation mode)."""
    SaleItemFormSet = formset_factory(SaleItemForm, extra=1, can_delete=True)
    if request.method == 'POST':
        sale_form = SaleForm(request.POST)
        item_formset = SaleItemFormSet(request.POST, prefix='items')
        payment_form = SalePaymentForm(request.POST, prefix='pay')
        valid = True
        if not sale_form.is_valid(): valid = False
        if not item_formset.is_valid(): valid = False
        if not payment_form.is_valid():
            if request.POST.get('pay-amount'):
                payment_form.add_error('amount', 'Enter a valid payment amount.')
                valid = False
        cleaned_items = []
        if item_formset.is_valid():
            for f in item_formset.forms:
                if f.cleaned_data.get('DELETE'): continue
                cd = f.cleaned_data
                if not cd.get('item_type') and not cd.get('inventory_item') and not cd.get('description'): continue
                if not cd.get('item_type'):
                    f.add_error('item_type', 'Select type'); valid = False; continue
                if cd.get('item_type') == 'non_inventory' and not cd.get('description'):
                    f.add_error('description', 'Description required'); valid = False; continue
                if cd.get('item_type') == 'inventory' and not cd.get('inventory_item'):
                    f.add_error('inventory_item', 'Choose inventory item'); valid = False; continue
                cleaned_items.append(cd)
        if not cleaned_items:
            messages.error(request, 'Add at least one valid item.'); valid = False
        from decimal import Decimal as _D
        total_amount = _D('0')
        for cd in cleaned_items:
            unit_price = cd.get('unit_price') or _D('0')
            if cd.get('item_type') == 'inventory' and cd.get('inventory_item') and (unit_price is None or unit_price <= 0):
                unit_price = _D(str(cd['inventory_item'].unit_price or 0))
            qty = cd.get('quantity') or _D('0')
            total_amount += _D(str(unit_price)) * _D(str(qty))
        payment_amount = None
        if payment_form.is_valid():
            payment_amount = payment_form.cleaned_data.get('amount')
            if payment_amount is not None and payment_amount < 0:
                payment_form.add_error('amount', 'Payment cannot be negative.'); valid = False
            elif payment_amount and payment_amount > total_amount:
                payment_form.add_error('amount', 'Payment exceeds total.'); valid = False
        if valid:
            from django.db import transaction
            with transaction.atomic():
                sale = sale_form.save(commit=False)
                sale.created_by = request.user
                sale.status = 'quote'
                sale.save()
                for cd in cleaned_items:
                    unit_price = cd.get('unit_price') or 0
                    if cd.get('item_type') == 'inventory' and cd.get('inventory_item') and (unit_price is None or unit_price <= 0):
                        unit_price = cd['inventory_item'].unit_price or 0
                    SaleItem.objects.create(
                        sale=sale,
                        item_type=cd['item_type'],
                        inventory_item=cd.get('inventory_item') if cd['item_type'] == 'inventory' else None,
                        description=(cd.get('description') or '') if cd['item_type'] != 'inventory' else (
                            f"{cd['inventory_item'].part_name} ({cd['inventory_item'].part_code})" if cd.get('inventory_item') else ''
                        ),
                        quantity=cd.get('quantity') or 0,
                        boxes=cd.get('boxes') or 0,
                        unit_price=unit_price,
                    )
                sale.recalc_total(save=True)
                # Usually quotations may not record payments; allow if provided
                payment = None
                if payment_amount and payment_amount > 0:
                    payment = payment_form.save(commit=False)
                    payment.sale = sale
                    payment.save()
                messages.success(request, 'Quotation created.')
                if payment:
                    messages.success(request, f'Payment recorded (Receipt {payment.receipt_number}).')
                return redirect('sale_detail', pk=sale.pk)
        else:
            parts = _collect_form_errors(sale_form=sale_form, item_formset=item_formset, payment_form=payment_form)
            messages.error(request, " | ".join(parts) if parts else 'Please correct the highlighted fields.')
    else:
        sale_form = SaleForm()
        item_formset = formset_factory(SaleItemForm, extra=1, can_delete=True)(prefix='items')
        payment_form = SalePaymentForm(prefix='pay')
    inventory_items = InventoryItem.objects.all()
    inv_prices = {str(i.id): float(i.unit_price or 0) for i in inventory_items}
    return render(request, 'core/sale_create_unified.html', {
        'sale_form': sale_form,
        'item_formset': item_formset,
        'payment_form': payment_form,
        'inventory_items': inventory_items,
        'inventory_prices': inv_prices,
        'title': 'Create Quotation'
    })


@login_required
@permission_required('core.add_customer', raise_exception=True)
def customer_quick_add(request):
    """AJAX endpoint to quickly add a customer and return JSON for select update."""
    if request.method != 'POST' or request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)
    form = CustomerForm(request.POST)
    if form.is_valid():
        cust = form.save()
        return JsonResponse({
            'id': cust.pk,
            'name': cust.name,
            'customer_id': cust.customer_id,
            'display': f"{cust.name} ({cust.customer_id})"
        })
    return JsonResponse({'errors': form.errors}, status=422)


@login_required
@permission_required('core.change_sale', raise_exception=True)
def sale_convert_to_invoice(request, pk):
    sale = _get_visible_sale_or_404(request, pk)
    if sale.status != 'quote':
        messages.info(request, 'This sale is not a quotation.')
        return redirect('sale_detail', pk=sale.pk)
    sale.status = 'draft'
    sale.save(update_fields=['status', 'updated_at'])
    messages.success(request, 'Quotation converted to invoice. You can now finalize when ready.')
    return redirect('sale_detail', pk=sale.pk)


@login_required
@permission_required('core.view_sale', raise_exception=True)
def sale_detail(request, pk):
    sale = _get_visible_sale_or_404(request, pk)
    items = sale.items.select_related('inventory_item').all()
    payments = sale.payments.all()
    add_item_form = None
    add_payment_form = None
    # Allow admins to edit finalized sales, or allow regular users to edit drafts
    can_edit_items = request.user.has_perm('core.add_saleitem') and (sale.status == 'draft' or request.user.is_superuser or request.user.is_staff)
    if can_edit_items:
        add_item_form = SaleItemForm()
    if request.user.has_perm('core.add_salepayment') and sale.status != 'cancelled' and sale.status != 'quote':
        add_payment_form = SalePaymentForm()
    # Inventory prices for client-side autofill in add-item form
    inv_qs = None
    if add_item_form:
        try:
            inv_qs = add_item_form.fields['inventory_item'].queryset
        except Exception:
            logger.exception('Error retrieving inventory_item queryset')
            inv_qs = InventoryItem.objects.none()
    inventory_prices = {str(i.id): float(i.unit_price) for i in (inv_qs or [])}
    context = {
        'sale': sale,
        'items': items,
        'payments': payments,
        'add_item_form': add_item_form,
        'add_payment_form': add_payment_form,
        'inventory_prices': inventory_prices,
    }
    return render(request, 'core/sale_detail.html', context)


@login_required
@permission_required('core.view_sale', raise_exception=True)
def sale_invoice(request, pk):
    sale = _get_visible_sale_or_404(request, pk)
    items = sale.items.select_related('inventory_item').all()
    context = {
        'sale': sale,
        'items': items,
    }
    # For quotations, calculate valid_until date (30 days from creation)
    if sale.status == 'quote':
        from datetime import timedelta
        context['valid_until'] = sale.created_at + timedelta(days=30)
    # Render quotation template for quotes, invoice template for all others
    template = 'core/sale_quotation.html' if sale.status == 'quote' else 'core/sale_invoice.html'
    return render(request, template, context)


@login_required
@permission_required('core.add_saleitem', raise_exception=True)
def sale_add_item(request, pk):
    sale = _get_visible_sale_or_404(request, pk)
    # Allow admins to edit finalized sales
    if sale.status != 'draft' and not (request.user.is_superuser or request.user.is_staff):
        messages.warning(request, 'Cannot add items to a finalized sale.')
        return redirect('sale_detail', pk=sale.pk)
    
    if request.method == 'POST':
        form = SaleItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.sale = sale
            # Basic validation: require description for non-inventory items
            if item.item_type == 'non_inventory' and not item.description:
                messages.error(request, 'Description is required for non-inventory items.')
            else:
                # Allow override: use provided unit_price if > 0 else fallback to inventory price
                if item.item_type == 'inventory' and item.inventory_item and (item.unit_price is None or item.unit_price <= 0):
                    item.unit_price = item.inventory_item.unit_price
                
                # If adding to a finalized sale, check and adjust inventory
                if sale.status == 'finalized' and item.item_type == 'inventory' and item.inventory_item:
                    inv = item.inventory_item
                    
                    # Validate boxes if provided
                    if (item.boxes or 0) > 0:
                        if inv.box_count < item.boxes:
                            messages.error(request, f"Insufficient box stock for {inv.part_name} ({inv.part_code}). Available boxes: {inv.box_count}, required: {item.boxes}")
                            return redirect('sale_detail', pk=sale.pk)
                    
                    # Validate unit quantity
                    if inv.quantity < item.quantity:
                        messages.error(request, f"Insufficient unit stock for {inv.part_name} ({inv.part_code}). Available: {inv.quantity}, required: {item.quantity}")
                        return redirect('sale_detail', pk=sale.pk)
                    
                    # Save the item first
                    item.save()
                    
                    # Deduct boxes if provided
                    if (item.boxes or 0) > 0:
                        prev_boxes = inv.box_count
                        inv.box_count = prev_boxes - item.boxes
                        inv.save(update_fields=['box_count', 'updated_at'])
                        
                        StockHistory.objects.create(
                            item=inv,
                            transaction_type='out',
                            quantity=0,
                            previous_quantity=0,
                            new_quantity=0,
                            box_quantity=item.boxes,
                            previous_box_quantity=prev_boxes,
                            new_box_quantity=inv.box_count,
                            reason=f"Added to finalized sale {sale.sale_number} by admin (boxes)",
                            created_by=request.user.username
                        )
                    
                    # Deduct unit quantity
                    previous = inv.quantity
                    inv.quantity = previous - item.quantity
                    inv.save(update_fields=['quantity', 'updated_at'])
                    
                    StockHistory.objects.create(
                        item=inv,
                        transaction_type='out',
                        quantity=item.quantity,
                        previous_quantity=previous,
                        new_quantity=inv.quantity,
                        box_quantity=0,
                        previous_box_quantity=0,
                        new_box_quantity=0,
                        reason=f"Added to finalized sale {sale.sale_number} by admin",
                        created_by=request.user.username
                    )
                    
                    messages.success(request, f'Item added to finalized sale and inventory adjusted: {inv.part_name} ({inv.part_code})')
                else:
                    item.save()
                    messages.success(request, 'Item added to sale.')
                
                # Recalculate sale total
                try:
                    sale.recalc_total(save=True)
                except Exception:
                    logger.exception('Failed to recalc total after adding item to Sale pk=%s', pk)
                
                return redirect('sale_detail', pk=sale.pk)
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"{field}: {err}")
    return redirect('sale_detail', pk=sale.pk)


@login_required
@permission_required('core.finalize_sale', raise_exception=True)
@require_POST
def sale_finalize(request, pk):
    sale = _get_visible_sale_or_404(request, pk)
    if sale.status == 'finalized':
        messages.info(request, 'Sale already finalized.')
        return redirect('sale_detail', pk=sale.pk)
    try:
        low_stock = sale.finalize(user=request.user)
        if low_stock:
            names = ', '.join([f"{i.part_name} ({i.part_code})" for i in low_stock])
            messages.warning(request, f"Finalized. Low stock: {names}")
        else:
            messages.success(request, 'Sale finalized successfully.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('sale_detail', pk=sale.pk)


@login_required
@permission_required('core.delete_saleitem', raise_exception=True)
def sale_delete_item(request, pk, item_pk):
    sale = _get_visible_sale_or_404(request, pk)
    # Allow admins to edit finalized sales
    if sale.status != 'draft' and not (request.user.is_superuser or request.user.is_staff):
        messages.warning(request, 'Cannot delete items from a finalized sale.')
        return redirect('sale_detail', pk=sale.pk)
    
    item = get_object_or_404(SaleItem, pk=item_pk, sale=sale)
    
    if request.method == 'POST':
        # If deleting from a finalized sale, restore inventory
        if sale.status == 'finalized' and item.item_type == 'inventory' and item.inventory_item:
            inv = item.inventory_item
            
            # Restore boxes if they were deducted
            if (item.boxes or 0) > 0:
                prev_boxes = inv.box_count
                inv.box_count = prev_boxes + item.boxes
                inv.save(update_fields=['box_count', 'updated_at'])
                
                StockHistory.objects.create(
                    item=inv,
                    transaction_type='adjustment',
                    quantity=0,
                    previous_quantity=0,
                    new_quantity=0,
                    box_quantity=item.boxes,
                    previous_box_quantity=prev_boxes,
                    new_box_quantity=inv.box_count,
                    reason=f"Reversed: Item deleted from finalized sale {sale.sale_number} by admin",
                    created_by=request.user.username
                )
            
            # Restore unit quantity
            previous = inv.quantity
            inv.quantity = previous + item.quantity
            inv.save(update_fields=['quantity', 'updated_at'])
            
            StockHistory.objects.create(
                item=inv,
                transaction_type='adjustment',
                quantity=item.quantity,
                previous_quantity=previous,
                new_quantity=inv.quantity,
                box_quantity=0,
                previous_box_quantity=0,
                new_box_quantity=0,
                reason=f"Reversed: Item deleted from finalized sale {sale.sale_number} by admin",
                created_by=request.user.username
            )
            
            messages.success(request, f'Item deleted and inventory restored: {inv.part_name} ({inv.part_code})')
        
        item.delete()
        try:
            sale.recalc_total(save=True)
        except Exception:
            logger.exception('Failed to recalc total after deleting item from Sale pk=%s', pk)
    return redirect('sale_detail', pk=sale.pk)


@login_required
@permission_required('core.view_salepayment', raise_exception=True)
def sale_payments_export(request, pk):
    import csv
    from django.utils.encoding import smart_str
    sale = _get_visible_sale_or_404(request, pk)
    payments = sale.payments.all().order_by('-payment_date', '-created_at')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{sale.sale_number}_payments.csv"'
    writer = csv.writer(response)
    writer.writerow(['Receipt', 'Date', 'Method', 'Amount'])
    for p in payments:
        writer.writerow([
            smart_str(p.receipt_number),
            p.payment_date,
            smart_str(p.get_method_display()),
            f"{p.amount}",
        ])
    return response


@login_required
@permission_required('core.view_salepayment', raise_exception=True)
def sale_payments_export_pdf(request, pk):
    """Export payment history for a sale as a professional PDF statement."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO
    from xml.sax.saxutils import escape as xml_escape
    from django.conf import settings
    from django.utils import timezone

    sale = _get_visible_sale_or_404(request, pk)
    payments = sale.payments.all().order_by('payment_date', 'created_at')

    # Brand info from settings
    brand_name = getattr(settings, 'BRAND_NAME', 'Organization')
    brand_address = getattr(settings, 'BRAND_ADDRESS', '')
    brand_phone = getattr(settings, 'BRAND_PHONE', '')
    brand_email = getattr(settings, 'BRAND_EMAIL', '')

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        topMargin=2*cm, 
        bottomMargin=2*cm, 
        leftMargin=2*cm, 
        rightMargin=2*cm
    )
    
    page_width = A4[0] - 4*cm  # Available width
    styles = getSampleStyleSheet()

    # Modern color palette - clean and professional
    accent_color = colors.HexColor('#2563EB')  # Modern blue accent
    text_primary = colors.HexColor('#111827')  # Deep black
    text_secondary = colors.HexColor('#6B7280')  # Medium gray
    text_light = colors.HexColor('#9CA3AF')  # Light gray
    divider_color = colors.HexColor('#E5E7EB')  # Light divider
    success_color = colors.HexColor('#10B981')  # Green
    danger_color = colors.HexColor('#EF4444')  # Red
    bg_subtle = colors.HexColor('#F9FAFB')  # Very light gray

    # Custom styles - modern typography
    styles.add(ParagraphStyle(
        name='BrandName',
        fontSize=20,
        textColor=text_primary,
        fontName='Helvetica-Bold',
        leading=24,
        spaceAfter=4,
        alignment=TA_LEFT
    ))
    styles.add(ParagraphStyle(
        name='BrandInfo',
        fontSize=9,
        textColor=text_secondary,
        fontName='Helvetica',
        leading=13,
        alignment=TA_LEFT
    ))
    styles.add(ParagraphStyle(
        name='DocTitle',
        fontSize=11,
        textColor=text_secondary,
        fontName='Helvetica',
        spaceAfter=2,
        alignment=TA_RIGHT,
        textTransform='uppercase'
    ))
    styles.add(ParagraphStyle(
        name='DocNumber',
        fontSize=16,
        textColor=text_primary,
        fontName='Helvetica-Bold',
        leading=20,
        alignment=TA_RIGHT
    ))
    styles.add(ParagraphStyle(
        name='DocDate',
        fontSize=9,
        textColor=text_secondary,
        fontName='Helvetica',
        alignment=TA_RIGHT
    ))
    styles.add(ParagraphStyle(
        name='SectionLabel',
        fontSize=8,
        textColor=text_light,
        fontName='Helvetica-Bold',
        spaceBefore=12,
        spaceAfter=6,
        textTransform='uppercase',
        letterSpacing=0.5
    ))
    styles.add(ParagraphStyle(
        name='CustomerName',
        fontSize=14,
        textColor=text_primary,
        fontName='Helvetica-Bold',
        leading=18,
        alignment=TA_LEFT
    ))
    styles.add(ParagraphStyle(
        name='InfoText',
        fontSize=9,
        textColor=text_secondary,
        fontName='Helvetica',
        leading=14,
        alignment=TA_LEFT
    ))
    styles.add(ParagraphStyle(
        name='TableHeader',
        fontSize=8,
        textColor=text_secondary,
        fontName='Helvetica-Bold',
        leading=11,
        textTransform='uppercase'
    ))
    styles.add(ParagraphStyle(
        name='TableCell',
        fontSize=9,
        textColor=text_primary,
        fontName='Helvetica',
        leading=13
    ))
    styles.add(ParagraphStyle(
        name='TableCellBold',
        fontSize=9,
        textColor=text_primary,
        fontName='Helvetica-Bold',
        leading=13
    ))
    styles.add(ParagraphStyle(
        name='TotalLabel',
        fontSize=10,
        textColor=text_secondary,
        fontName='Helvetica',
        leading=14
    ))
    styles.add(ParagraphStyle(
        name='TotalValue',
        fontSize=13,
        textColor=text_primary,
        fontName='Helvetica-Bold',
        leading=16,
        alignment=TA_RIGHT
    ))
    styles.add(ParagraphStyle(
        name='FooterNote',
        fontSize=8,
        textColor=text_light,
        fontName='Helvetica',
        leading=12,
        alignment=TA_LEFT
    ))

    elements = []

    def fmt_currency(val):
        """Format currency without symbol"""
        return f"{float(val):,.2f}"

    # ============== HEADER SECTION ==============
    # Two-column header: Brand on left, Document info on right
    header_left = []
    header_left.append(Paragraph(xml_escape(brand_name), styles['BrandName']))
    if brand_address:
        header_left.append(Paragraph(xml_escape(brand_address), styles['BrandInfo']))
    contact_line = []
    if brand_phone:
        contact_line.append(f"{brand_phone}")
    if brand_email:
        contact_line.append(f"{brand_email}")
    if contact_line:
        header_left.append(Paragraph(" • ".join(contact_line), styles['BrandInfo']))
    
    header_right = []
    header_right.append(Paragraph("PAYMENT STATEMENT", styles['DocTitle']))
    header_right.append(Paragraph(f"#{sale.sale_number}", styles['DocNumber']))
    header_right.append(Paragraph(timezone.now().strftime('%B %d, %Y'), styles['DocDate']))
    
    header_table = Table([
        [
            [p for p in header_left],
            [p for p in header_right]
        ]
    ], colWidths=[page_width * 0.60, page_width * 0.40])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.6*cm))
    elements.append(HRFlowable(width="100%", thickness=1, color=divider_color, spaceAfter=0.6*cm))

    # ============== BILL TO SECTION ==============
    customer_name = sale.customer.name if sale.customer else 'N/A'
    customer_phone = getattr(sale.customer, 'phone', '') if sale.customer else ''
    customer_address = getattr(sale.customer, 'address', '') if sale.customer else ''
    
    elements.append(Paragraph("BILL TO", styles['SectionLabel']))
    elements.append(Paragraph(xml_escape(customer_name), styles['CustomerName']))
    if customer_phone:
        elements.append(Paragraph(f"Phone: {customer_phone}", styles['InfoText']))
    if customer_address:
        elements.append(Paragraph(xml_escape(customer_address), styles['InfoText']))
    
    elements.append(Spacer(1, 0.6*cm))

    # ============== ORDER SUMMARY SECTION ==============
    elements.append(Paragraph("ORDER SUMMARY", styles['SectionLabel']))
    
    # Get product names from sale items
    sale_items = sale.items.all()
    product_names = []
    for item in sale_items:
        if item.item_type == 'inventory' and item.inventory_item:
            product_names.append(item.inventory_item.part_name)
        elif item.description:
            product_names.append(item.description[:60])
    product_name_str = ", ".join(product_names) if product_names else "N/A"
    
    order_summary_data = [
        [
            Paragraph("PRODUCT", styles['TableHeader']),
            Paragraph("ORDER DATE", styles['TableHeader']),
            Paragraph("STATUS", styles['TableHeader']),
            Paragraph("AMOUNT", styles['TableHeader']),
        ],
        [
            Paragraph(xml_escape(product_name_str), styles['TableCell']),
            Paragraph(sale.created_at.strftime('%Y-%m-%d') if sale.created_at else '-', styles['TableCell']),
            Paragraph(sale.get_status_display(), styles['TableCellBold']),
            Paragraph(fmt_currency(sale.total_amount), styles['TableCellBold']),
        ],
    ]
    
    order_table = Table(order_summary_data, colWidths=[
        page_width * 0.40,
        page_width * 0.20,
        page_width * 0.20,
        page_width * 0.20
    ])
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), bg_subtle),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, 1), 12),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('LINEBELOW', (0, 0), (-1, 0), 1, divider_color),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(order_table)
    elements.append(Spacer(1, 0.6*cm))

    # ============== PAYMENT HISTORY TABLE ==============
    elements.append(Paragraph("PAYMENT HISTORY", styles['SectionLabel']))

    if payments.exists():
        payment_data = [[
            Paragraph("RECEIPT", styles['TableHeader']),
            Paragraph("DATE", styles['TableHeader']),
            Paragraph("METHOD", styles['TableHeader']),
            Paragraph("NOTES", styles['TableHeader']),
            Paragraph("AMOUNT", styles['TableHeader']),
        ]]
        
        for p in payments:
            notes_text = (p.notes[:45] + '...' if len(p.notes) > 45 else p.notes) if p.notes else '-'
            payment_data.append([
                Paragraph(p.receipt_number, styles['TableCell']),
                Paragraph(str(p.payment_date), styles['TableCell']),
                Paragraph(p.get_method_display(), styles['TableCell']),
                Paragraph(xml_escape(notes_text), styles['TableCell']),
                Paragraph(fmt_currency(p.amount), styles['TableCellBold']),
            ])

        payments_table = Table(payment_data, colWidths=[
            page_width * 0.20,
            page_width * 0.16,
            page_width * 0.14,
            page_width * 0.32,
            page_width * 0.18
        ])
        
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), bg_subtle),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('LINEBELOW', (0, 0), (-1, 0), 1, divider_color),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        
        # Subtle dividers between rows
        for i in range(1, len(payment_data) - 1):
            table_style.append(('LINEBELOW', (0, i), (-1, i), 0.5, divider_color))
        
        payments_table.setStyle(TableStyle(table_style))
        elements.append(payments_table)
    else:
        elements.append(Paragraph("No payments recorded for this order.", styles['InfoText']))

    elements.append(Spacer(1, 0.8*cm))

    # ============== TOTALS SUMMARY ==============
    totals_data = [
        [
            Paragraph("Total Amount", styles['TotalLabel']),
            Paragraph(fmt_currency(sale.total_amount), styles['TotalValue']),
        ],
        [
            Paragraph("Total Paid", styles['TotalLabel']),
            Paragraph(f'<font color="{success_color}">{fmt_currency(sale.total_paid)}</font>', styles['TotalValue']),
        ],
    ]
    
    # Balance due row with color coding
    balance_color = danger_color if sale.balance_due > 0 else success_color
    totals_data.append([
        Paragraph("<b>Balance Due</b>", styles['TotalLabel']),
        Paragraph(f'<font color="{balance_color}"><b>{fmt_currency(sale.balance_due)}</b></font>', styles['TotalValue']),
    ])
    
    totals_table = Table(totals_data, colWidths=[page_width * 0.50, page_width * 0.50])
    totals_table.setStyle(TableStyle([
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('LINEABOVE', (0, 0), (-1, 0), 1, divider_color),
        ('LINEABOVE', (0, -1), (-1, -1), 2, text_primary),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(totals_table)

    # ============== FOOTER ==============
    elements.append(Spacer(1, 0.8*cm))
    elements.append(HRFlowable(width="100%", thickness=1, color=divider_color, spaceAfter=0.4*cm))
    footer_note = f"This is a computer-generated document. Generated on {timezone.now().strftime('%B %d, %Y at %H:%M')}."
    elements.append(Paragraph(footer_note, styles['FooterNote']))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{sale.sale_number}_payment_statement.pdf"'
    return response


@login_required
@permission_required('core.add_salepayment', raise_exception=True)
def sale_add_payment(request, pk):
    sale = _get_visible_sale_or_404(request, pk)
    if sale.status == 'cancelled':
        messages.warning(request, 'Cannot record payments for a cancelled sale.')
        return redirect('sale_detail', pk=sale.pk)

    if request.method == 'POST':
        form = SalePaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            # Prevent overpayment beyond balance due
            if payment.amount <= 0:
                messages.error(request, 'Amount must be greater than zero.')
            else:
                with transaction.atomic():
                    locked_sale = _visible_sales_queryset(request.user).select_for_update().get(pk=sale.pk)
                    Customer.objects.select_for_update().get(pk=locked_sale.customer_id)

                    if payment.amount > locked_sale.balance_due:
                        messages.error(request, 'Payment exceeds remaining balance.')
                    else:
                        payment.sale = locked_sale
                        payment.save()
                        # Log to ledger as a credit (inflow)
                        try:
                            LedgerEntry.objects.update_or_create(
                                source='sale_payment',
                                reference=payment.receipt_number,
                                defaults={
                                    'entry_type': 'credit',
                                    'description': f"Payment for {locked_sale.sale_number}",
                                    'amount': payment.amount,
                                }
                            )
                        except Exception:
                            logger.exception('Non-blocking ledger write failure for SalePayment receipt=%s', payment.receipt_number)
                        messages.success(request, f'Payment recorded. Receipt: {payment.receipt_number}')
                        return redirect('sale_detail', pk=locked_sale.pk)
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"{field}: {err}")
    return redirect('sale_detail', pk=sale.pk)


@login_required
@permission_required('core.change_salepayment', raise_exception=True)
def sale_edit_payment(request, pk, payment_pk):
    sale = _get_visible_sale_or_404(request, pk)
    payment = get_object_or_404(SalePayment, pk=payment_pk, sale=sale)

    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Only admin users can edit sale payments.')
        return redirect('sale_detail', pk=sale.pk)

    if request.method == 'POST':
        form = SalePaymentForm(request.POST, instance=payment)
        if form.is_valid():
            updated_payment = form.save(commit=False)

            if updated_payment.amount <= 0:
                messages.error(request, 'Amount must be greater than zero.')
                return redirect('sale_edit_payment', pk=sale.pk, payment_pk=payment.pk)

            with transaction.atomic():
                locked_sale = _visible_sales_queryset(request.user).select_for_update().get(pk=sale.pk)
                Customer.objects.select_for_update().get(pk=locked_sale.customer_id)
                locked_payment = get_object_or_404(SalePayment.objects.select_for_update(), pk=payment.pk, sale=locked_sale)

                other_paid = locked_sale.payments.exclude(pk=locked_payment.pk).aggregate(total=Sum('amount')).get('total') or 0
                max_allowed = locked_sale.total_amount - other_paid
                if updated_payment.amount > max_allowed:
                    messages.error(request, 'Payment exceeds remaining balance.')
                    return redirect('sale_edit_payment', pk=locked_sale.pk, payment_pk=locked_payment.pk)

                locked_payment.amount = updated_payment.amount
                locked_payment.payment_date = updated_payment.payment_date
                locked_payment.method = updated_payment.method
                locked_payment.notes = updated_payment.notes
                locked_payment.save()

                try:
                    LedgerEntry.objects.update_or_create(
                        source='sale_payment',
                        reference=locked_payment.receipt_number,
                        defaults={
                            'entry_type': 'credit',
                            'description': f"Payment for {locked_sale.sale_number}",
                            'amount': locked_payment.amount,
                        }
                    )
                except Exception:
                    logger.exception('Non-blocking ledger update failure for SalePayment receipt=%s', locked_payment.receipt_number)

                messages.success(request, f'Payment {locked_payment.receipt_number} updated successfully.')
                return redirect('sale_detail', pk=locked_sale.pk)
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"{field}: {err}")
    else:
        form = SalePaymentForm(instance=payment)

    context = {
        'sale': sale,
        'payment': payment,
        'form': form,
    }
    return render(request, 'core/sale_payment_edit.html', context)


@login_required
@permission_required('core.delete_salepayment', raise_exception=True)
@require_POST
def sale_delete_payment(request, pk, payment_pk):
    sale = _get_visible_sale_or_404(request, pk)
    payment = get_object_or_404(SalePayment, pk=payment_pk, sale=sale)

    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Only admin users can delete sale payments.')
        return redirect('sale_detail', pk=sale.pk)

    with transaction.atomic():
        locked_sale = _visible_sales_queryset(request.user).select_for_update().get(pk=sale.pk)
        Customer.objects.select_for_update().get(pk=locked_sale.customer_id)
        locked_payment = get_object_or_404(SalePayment.objects.select_for_update(), pk=payment.pk, sale=locked_sale)
        receipt_number = locked_payment.receipt_number
        locked_payment.delete()

    try:
        LedgerEntry.objects.filter(source='sale_payment', reference=receipt_number).delete()
    except Exception:
        logger.exception('Non-blocking ledger cleanup failure for SalePayment receipt=%s', receipt_number)

    messages.success(request, f'Payment {receipt_number} deleted successfully.')
    return redirect('sale_detail', pk=sale.pk)


@login_required
@permission_required('core.view_salepayment', raise_exception=True)
def sale_payment_receipt(request, sale_pk, payment_pk):
    sale = _get_visible_sale_or_404(request, sale_pk)
    payment = get_object_or_404(SalePayment, pk=payment_pk, sale=sale)
    context = {
        'sale': sale,
        'payment': payment,
    }
    return render(request, 'core/sale_payment_receipt.html', context)


@login_required
@permission_required('core.view_sale', raise_exception=True)
def sales_export_csv(request):
    """Export sales/orders history as CSV, limited to Orders tab semantics (finalized only)."""
    import csv
    from django.utils.encoding import smart_str
    # Optional filters similar to sale_list
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    item_type = request.GET.get('item_type', '')
    qs = _visible_sales_queryset(request.user).filter(status='finalized').order_by('-created_at')
    if query:
        qs = qs.filter(Q(sale_number__icontains=query) | Q(customer__name__icontains=query))
    if status:
        qs = qs.filter(status=status)
    if item_type:
        qs = qs.filter(items__item_type=item_type).distinct()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sales_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['Sale Number', 'Customer', 'Status', 'Created At', 'Total', 'Paid', 'Due'])
    for s in qs:
        writer.writerow([
            smart_str(s.sale_number),
            smart_str(getattr(s.customer, 'name', '')),
            smart_str(s.status),
            s.created_at,
            f"{s.total_amount}",
            f"{s.total_paid}",
            f"{s.balance_due}",
        ])
    return response


@login_required
@permission_required('core.view_sale', raise_exception=True)
def sales_export_pdf(request):
    """Export orders (finalized sales) as a professional PDF built from HTML.
    Uses a dedicated template with branding, layout, and typography.
    """

    query = request.GET.get('q', '')
    status = 'finalized'  # enforce Orders tab semantics
    item_type = request.GET.get('item_type', '')
    customer_id = request.GET.get('customer_id')

    qs = _visible_sales_queryset(request.user).prefetch_related('items__inventory_item').filter(status=status).order_by('-created_at')
    if customer_id:
        qs = qs.filter(customer_id=customer_id)
    if query:
        qs = qs.filter(Q(sale_number__icontains=query) | Q(customer__name__icontains=query))
    if item_type:
        qs = qs.filter(items__item_type=item_type).distinct()

    sales = list(qs)
    has_orders = len(sales) > 0

    # Compute totals
    total_total = sum((float(s.total_amount or 0) for s in sales), start=0.0)
    total_paid = sum((float(s.total_paid or 0) for s in sales), start=0.0)
    total_due = sum((float(s.balance_due or 0) for s in sales), start=0.0)

    # Build Executive/Financial Report style PDF
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm, mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO
    from xml.sax.saxutils import escape as xml_escape

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    
    # Executive report styles with serif fonts
    styles.add(ParagraphStyle(
        name='CompanyHeader', 
        fontSize=10, 
        textColor=colors.HexColor('#334155'),
        fontName='Times-Roman',
        spaceAfter=2,
        alignment=TA_LEFT
    ))
    styles.add(ParagraphStyle(
        name='ReportTitle', 
        fontSize=24, 
        textColor=colors.HexColor('#1E293B'), 
        fontName='Times-Bold', 
        spaceAfter=8,
        leading=28,
        alignment=TA_LEFT,
        leftIndent=0
    ))
    styles.add(ParagraphStyle(
        name='ReportSubtitle', 
        fontSize=11, 
        textColor=colors.HexColor('#64748B'),
        fontName='Times-Roman',
        spaceAfter=2
    ))
    styles.add(ParagraphStyle(
        name='SummaryLabel', 
        fontSize=10, 
        textColor=colors.HexColor('#64748B'),
        fontName='Helvetica',
        leading=12
    ))
    styles.add(ParagraphStyle(
        name='SummaryValue', 
        fontSize=26, 
        textColor=colors.HexColor('#0F172A'), 
        fontName='Helvetica-Bold',
        leading=30
    ))

    elements = []

    def fmt_currency(val):
        # Keep plain numeric values to avoid unsupported glyph rendering blocks in PDF cells.
        return f"{float(val):,.2f}"

    def fmt_quantity(val):
        qty = float(val or 0)
        if qty.is_integer():
            return f"{int(qty)}"
        return f"{qty:.3f}".rstrip('0').rstrip('.')

    text_cell_style = ParagraphStyle(
        name='TextCell',
        fontSize=9,
        leading=11,
        fontName='Helvetica',
        alignment=TA_LEFT,
        wordWrap='CJK',
    )

    money_cell_style = ParagraphStyle(
        name='MoneyCell',
        fontSize=8,
        leading=9,
        fontName='Helvetica',
        alignment=TA_RIGHT,
        # Do not wrap numbers; prefer single-line values.
        splitLongWords=False,
    )

    def text_cell(value) -> Paragraph:
        return Paragraph(xml_escape(str(value or '')), text_cell_style)

    def money_cell(value) -> Paragraph:
        return Paragraph(xml_escape(fmt_currency(value)), money_cell_style)

    # Professional header with company area
    customer_obj = Customer.objects.filter(pk=customer_id).first() if customer_id else None
    
    # Header box with company info
    from django.utils import timezone as django_tz
    local_now = django_tz.localtime(django_tz.now())
    header_data = [[
        Paragraph("<b>Fashion Express</b><br/><font size=8>Financial Report</font>", styles['CompanyHeader']),
        Paragraph(f"<font size=8>Report Date: {local_now.strftime('%B %d, %Y')}<br/>Generated: {local_now.strftime('%I:%M %p')}</font>", 
                 ParagraphStyle(name='HeaderRight', parent=styles['CompanyHeader'], alignment=TA_RIGHT))
    ]]
    header_table = Table(header_data, colWidths=[15*cm, 10*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.3*cm))
    
    # Title section
    elements.append(Paragraph("Customer Orders Report", styles['ReportTitle']))
    subtitle_parts = ["Finalized Sales"]
    if customer_obj:
        subtitle_parts.append(f"Customer: {customer_obj.name}")
    elements.append(Paragraph(" | ".join(subtitle_parts), styles['ReportSubtitle']))
    elements.append(Spacer(1, 0.5*cm))
    
    # Horizontal line separator
    line_table = Table([['']], colWidths=[25*cm])
    line_table.setStyle(TableStyle([
        ('LINEABOVE', (0,0), (-1,0), 2, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 0.6*cm))

    # Professional data table with explicit item details per order line.
    table_data = [['Date', 'Sale Number', 'Product Name', 'Quantity', 'Unit Price', 'Sub Total', 'Total', 'Paid', 'Due']]
    span_configs = []  # Track which cells to span across multiple rows
    
    for s in sales:
        sale_items = list(s.items.all())
        start_row = len(table_data)
        
        if not sale_items:
            table_data.append([
                s.created_at.strftime('%Y-%m-%d'),
                text_cell(s.sale_number),
                text_cell('-'),
                text_cell('0'),
                money_cell(0),
                money_cell(0),
                money_cell(s.total_amount),
                money_cell(s.total_paid),
                money_cell(s.balance_due),
            ])
            continue

        for idx, item in enumerate(sale_items):
            if item.item_type == 'inventory' and item.inventory_item:
                product_name = (item.inventory_item.part_name or '').strip()
            else:
                product_name = (item.description or '').strip()

            row = [
                s.created_at.strftime('%Y-%m-%d') if idx == 0 else '',
                text_cell(s.sale_number) if idx == 0 else '',
                text_cell(product_name or '-'),
                text_cell(fmt_quantity(item.quantity)),
                money_cell(item.unit_price),
                money_cell(item.line_total),
                money_cell(s.total_amount) if idx == 0 else '',
                money_cell(s.total_paid) if idx == 0 else '',
                money_cell(s.balance_due) if idx == 0 else '',
            ]
            table_data.append(row)
            
            # Record span config: (col_idx, start_row, end_row) for cell merging
            if idx == 0 and len(sale_items) > 1:
                span_configs.append((0, start_row, start_row + len(sale_items) - 1))  # Date column
                span_configs.append((1, start_row, start_row + len(sale_items) - 1))  # Sale Number column
                span_configs.append((6, start_row, start_row + len(sale_items) - 1))  # Total column
                span_configs.append((7, start_row, start_row + len(sale_items) - 1))  # Paid column
                span_configs.append((8, start_row, start_row + len(sale_items) - 1))  # Due column

    if len(table_data) == 1:
        table_data.append(['', '', 'No orders found', '', '', '', '', '', ''])

    total_row_idx = None
    if has_orders:
        table_data.append(['TOTAL', '', '', '', '', '', money_cell(total_total), money_cell(total_paid), money_cell(total_due)])
        total_row_idx = len(table_data) - 1

    # Column widths tuned for landscape A4 so labels and numeric values do not overlap.
    data_table = Table(table_data, colWidths=[2.2*cm, 3.2*cm, 5.5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.6*cm], repeatRows=1)
    data_table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#334155')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (2,-1), 'LEFT'),
        ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,0), 7),
        ('FONTSIZE', (0,1), (-1,-1), 8),

        # Thick borders
        ('BOX', (0,0), (-1,-1), 2, colors.HexColor('#334155')),
        ('LINEBELOW', (0,0), (-1,0), 2, colors.white),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),

        # Row backgrounds
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),

        # Padding (tighter header so labels don't clip)
        ('TOPPADDING', (0,0), (-1,0), 7),
        ('BOTTOMPADDING', (0,0), (-1,0), 7),
        ('LEFTPADDING', (0,0), (-1,0), 3),
        ('RIGHTPADDING', (0,0), (-1,0), 3),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('LEFTPADDING', (0,1), (-1,-1), 4),
        ('RIGHTPADDING', (0,1), (-1,-1), 4),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    
    # Apply cell spanning for Date and Sale Number columns across multiple items
    # Also add dividers between sale groups
    divider_rows = []
    for col_idx, start_row, end_row in span_configs:
        data_table.setStyle(TableStyle([
            ('SPAN', (col_idx, start_row), (col_idx, end_row)),
        ]))
        divider_rows.append(end_row)
    
    # Add divider lines after each sale group
    for divider_row in divider_rows:
        data_table.setStyle(TableStyle([
            ('LINEBELOW', (0, divider_row), (-1, divider_row), 1.5, colors.HexColor('#9CA3AF')),
        ]))

    if total_row_idx is not None:
        data_table.setStyle(TableStyle([
            ('SPAN', (0, total_row_idx), (5, total_row_idx)),
            ('ALIGN', (0, total_row_idx), (5, total_row_idx), 'RIGHT'),
            ('FONTNAME', (0, total_row_idx), (-1, total_row_idx), 'Helvetica-Bold'),
            ('BACKGROUND', (0, total_row_idx), (-1, total_row_idx), colors.HexColor('#E2E8F0')),
            ('LINEABOVE', (0, total_row_idx), (-1, total_row_idx), 1, colors.HexColor('#334155')),
        ]))
    elements.append(data_table)

    # Professional footer with confidentiality notice
    elements.append(Spacer(1, 1.2*cm))
    
    # Footer line
    footer_line = Table([['']], colWidths=[25*cm])
    footer_line.setStyle(TableStyle([
        ('LINEABOVE', (0,0), (-1,0), 1, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(footer_line)
    elements.append(Spacer(1, 0.3*cm))
    
    footer_style = ParagraphStyle(
        name='Footer', 
        fontSize=8, 
        textColor=colors.HexColor('#94A3B8'), 
        alignment=TA_CENTER,
        fontName='Times-Italic'
    )
    footer_text = f"<b>CONFIDENTIAL</b> · This report contains proprietary information · Fashion Express © {timezone.now().year}"
    elements.append(Paragraph(footer_text, footer_style))
    
    # Page number placeholder
    page_num_style = ParagraphStyle(
        name='PageNum', 
        fontSize=8, 
        textColor=colors.HexColor('#CBD5E1'), 
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph("Page 1", page_num_style))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    # Generate filename with customer name if available
    if customer_obj:
        safe_customer_name = customer_obj.name.replace(' ', '_').replace('/', '_')
        filename = f"orders_report_{safe_customer_name}.pdf"
    else:
        filename = "orders_report.pdf"

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf_bytes)
    return response



@login_required
@permission_required('core.delete_sale', raise_exception=True)
def sale_delete(request, pk):
    """Delete a sale safely.
    Policy: allow deletion only for draft sales (payments allowed).
    """
    sale = _get_visible_sale_or_404(request, pk)
    if request.method == 'POST':
        if sale.status != 'draft':
            messages.warning(request, 'Only draft sales can be deleted.')
            return redirect('sale_detail', pk=sale.pk)
        sale.delete()
        messages.success(request, 'Sale deleted successfully.')
        return redirect('sale_list')
    # Confirm page reuse generic confirm template
    return render(request, 'core/confirm_delete.html', {'object': sale, 'type': 'Sale'})

